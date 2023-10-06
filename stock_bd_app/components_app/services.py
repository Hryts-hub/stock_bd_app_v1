
from components_app.models import DEFAULT_NAN_VALUE
from tqdm import tqdm
import pandas as pd
import openpyxl
import io
import time
import psutil
from django.db import connection

import os
from dotenv import load_dotenv

load_dotenv()

# offset for indexes is manually set to correspond art_numbers
DEFAULT_OFFSET_FOR_INDEXES = 2


# FUNCTIONS ------------------------------------------

def clear_table_records(model_class):
    model_class.objects.all().delete()

# CLASSES --------------------------------------------


class DataReader:
    def __init__(self, path_to_file, file_name):
        self.path_to_file = path_to_file
        self.file_name = file_name
        self.path = os.path.abspath(os.path.join(self.path_to_file, self.file_name))
        self.data = None
        self.msg = ''

    def read_data_from_stock_file(self, sheet_name, cols):
        if self.path_to_file is not None and os.path.exists(self.path):
            try:
                self.data = pd.read_excel(
                    self.path,
                    sheet_name=sheet_name,
                    usecols=cols,
                    engine='openpyxl',
                    dtype='object',
                    # This data type allows the columns to contain a mix of strings, numbers,
                    # and NaN values without any automatic type conversions
                )
            except Exception as e:
                self.msg = f'ERROR: {type(e)}: {e}'
        else:
            self.msg = f'{self.path} -- not found.'
            print(self.msg)
        return self.data, self.msg

    def read_data_from_stock_file_by_openpyxl(self, sheet_name, cell_names_list):
        excel_file_column_names = dict()
        if self.path_to_file is not None and os.path.exists(self.path):
            try:
                # read-only mode optimized for faster performance
                # only interested in reading specific data
                # like column names and don't need to modify the workbook.

                # When you load an Excel file using openpyxl.load_workbook(file),
                # the file is not automatically closed.
                # You are responsible for explicitly closing it when you're done using it.
                # Failing to do so may lead to file locks or resource conflicts,
                # which could result in a PermissionError when trying to access the file later.

                # wb._archive.close() and wb.close() did not work in read_only mode
                # --> https://stackoverflow.com/questions/31416842/
                # /openpyxl-does-not-close-excel-workbook-in-read-only-mode

                # SOLUTION: Using a context manager ensures that the file is
                # properly closed when the block is exited.
                # The file is opened for binary reading because Excel files are binary files.

                with open(self.path, "rb") as f:
                    in_mem_file = io.BytesIO(f.read())
                    # file context stored in an in-memory binary stream using the io.BytesIO class

                wb = openpyxl.load_workbook(in_mem_file, read_only=True)
                sheet = wb[sheet_name]
                excel_file_column_names = {
                    cell_name: sheet[cell_name].value for cell_name in cell_names_list
                }
            except Exception as e:
                print(f'ERROR: {type(e)}: {e}')
        else:
            self.msg = f'{self.path} -- not found.'
            print(self.msg)
        return excel_file_column_names

    def read_comments_from_stock_file_by_openpyxl(self, sheet_name, cell_names_dict):
        comment_dict = dict()

        if self.path_to_file is not None and os.path.exists(self.path):
            try:
                wb = openpyxl.load_workbook(self.path)
                sheet = wb[sheet_name]
                for cell_col_name, cell_names_list in cell_names_dict.items():
                    cell_names_list = [
    sheet[cell_name].comment.text if sheet[cell_name].comment else None for cell_name in cell_names_list
                    ]
                    comment_dict[sheet[cell_col_name].value] = cell_names_list
                wb.close()
            except Exception as e:
                print(f'ERROR: {type(e)}: {e}')
        else:
            self.msg = f'{self.path} -- not found.'
            print(self.msg)
        return comment_dict


# -------------------- functions

# this func will be replaced by updated one.
# table should be filled from the source tables
# def create_stock_components():
#     # read stock file --> df_stock_components
#     cols = 'C, D, E, F, G'
#     sheet_name = 'Склад'
#     df_stock_components, msg = DataReader(
#         PATH_TO_FILE_STOCK, FILE_STOCK).read_data_from_stock_file(sheet_name, cols)
#
#     # print(df_stock_components.shape)
#     # print(df_stock_components.dtypes)
#
#     df_stock_components = df_stock_components.dropna(subset=['Артикул'])
#     df_stock_components = df_stock_components.astype({'Артикул': int})
#     # df_stock_components = df_stock_components.iloc[12:17]
#
#     created_count = 0
#     err_count = 0
#
#     # fill DB by df
#     for raw in tqdm(df_stock_components.itertuples(index=False)):  # index=False
#         try:
#             # print(raw)
#             # print(raw[0])
#             # print(raw[1])
#             # print(raw[4])
#             StockComponents.objects.create(
#                 art_number=raw[0],
#                 component_stock_naming=raw[4],
#                 category=raw[1],
#                 type=raw[2],
#                 subtype=raw[3],
#                 user=CustomUser.objects.get(username='hryts')  # 'hryts'
#             )
#             created_count += 1
#         except Exception as e:
#             print(e)
#             print(raw)
#             err_count += 1
#             pass
#
#     print(f'created_count = {created_count}, err_count = {err_count}')
#
#     pass


class StockTabFromExcelUpdater:
    def __init__(self, model, path_to_file, file_name, sheet_name, columns):
        self.model = model
        self.path_to_file = path_to_file
        self.file_name = file_name
        self.table_name = model._meta.db_table
        self.db_column_names = {
            field.name: field.verbose_name for field in self.model._meta.get_fields()
        }
        self.db_verbose_names = {
            field.verbose_name: field.name for field in self.model._meta.get_fields()
        }
        self.sheet_name = sheet_name
        self.columns = columns  # "A, B, C..."

        # read not all cols from file, but just self.columns using cell_names
        self.column_names_from_excel_file = DataReader(
            self.path_to_file, self.file_name
        ).read_data_from_stock_file_by_openpyxl(self.sheet_name, self.make_cell_names(self.columns))

        self.col_list = [
            k for k, v in self.db_column_names.items() if v in self.column_names_from_excel_file.values()
        ]

        self.nan_value = DEFAULT_NAN_VALUE
        self.offset = DEFAULT_OFFSET_FOR_INDEXES
        self.fields_of_comments_dict = self.find_comment_colum_names()

    @staticmethod
    def make_cell_names(columns):
        """
        :param columns: str /// 'A, B, C'
        :return: cell_names: list /// ['A1', 'B1', 'C1']
        """
        columns_list = columns.split(', ')
        cell_names = [f'{column_letter}1' for column_letter in columns_list]
        return cell_names

    def validate_columns(self):
        """
        Checks that excel_file_col_names (input cols for create-update, not all cols) IN DB_col_names
        If NOT -->> flag is False -->> need to alter table
        :return: flag
        """

        flag = True

        if self.column_names_from_excel_file:
            for cell, col in self.column_names_from_excel_file.items():
                if col not in self.db_column_names.values():
                    flag = False
                    print(f'ERR-CELL is << {cell} >> (column_name of this cell not found in DB): {col}')
        else:
            flag = False
        # print(flag)
        return flag

    def read_specified_columns_from_db_table(self, col_list):
        """
        Reads data from DB-table.
        Col_names filtered using verbose names of DB-columns,
        and index-numerating corrected with offset
        to get DF with the same structure as df_from_excel_file,
        and indexes that can be used in comparison of two DFs
        :return: DF with specified_columns_data from DB
        """

        columns_str = ', '.join(col_list)
        with connection.cursor() as cursor:
            query = f"SELECT {columns_str} FROM {self.table_name};"
            cursor.execute(query)
            data = cursor.fetchall()

        field_names = [self.db_column_names[field_name] for field_name in col_list]
        db_df = pd.DataFrame(data, columns=field_names)
        db_df.index = db_df.index + self.offset

        return db_df

    def read_specified_columns_from_excel_sheet(self, db_art_max):
        """
        Reads data from excel_file by DataReader class.
        Index-numerating corrected with offset,
        empty-tail of DF discarded using art_number_max,
        None values filled with default nan_value of str type,
        all values in DF converted to STR type.
        :return: DF
        """

        source_df, msg = DataReader(
            self.path_to_file, self.file_name).read_data_from_stock_file(
            self.sheet_name, self.columns)

        source_df.index = source_df.index + self.offset

        # if some last art_numbers were deleted from file - this changes updated in DB by default nan_value
        art_number_max = max(source_df['Артикул'].max(), db_art_max)
        source_df = source_df[:(art_number_max-1)]

        # all values in columns (except index) should be str - and nan_value is str type
        source_df = source_df.fillna(self.nan_value)

        type_dict = {col: str for col in self.column_names_from_excel_file.values()}
        source_df = source_df.astype(type_dict)

        return source_df

    def validate_df_to_get_unequal_rows(self):
        """
        Compares DF from DB and DF from excel file.
        First it returns rows with unequal_indexes to create this records in BD.
        Second it returns unequal_rows to update this records in BD.
        If column_names incorrect so self.validate_columns()=False -> (None, None) returned
        :return: DF, len(DF from excel file): int
        """
        if self.validate_columns():
            db_df = self.read_specified_columns_from_db_table(self.col_list)
            db_art_max = db_df.index.max()
            source_df = self.read_specified_columns_from_excel_sheet(db_art_max)

            # unequal_indexes is an array containing the indexes
            # that are differs between the two DataFrames.
            if not source_df.empty and not db_df.empty:
                unequal_indexes = source_df.index[~source_df.index.isin(db_df.index)].tolist()
            else:
                unequal_indexes = []

            # new rows are created first, then unequal rows are updated
            if unequal_indexes:
                res = source_df.loc[unequal_indexes]
                # return source_df.loc[unequal_indexes]
            elif db_df.empty or source_df.empty:
                res = source_df
                # return source_df
            else:
                # can't compare None (None --> default nan value = ''
                comparison_result = db_df == source_df
                # unequal_rows = source_df[~comparison_result.all(axis=1)]
                res = source_df[~comparison_result.all(axis=1)]

                # return unequal_rows
            return res, len(source_df)
        else:
            return None, None

    def create_by_iterrows(self, df):
        # FOR TEST PURPOSES

        created_count = 0
        err_count = 0

        start_time = time.time()
        mem_before = psutil.virtual_memory().used
        # col names in df are verbose names for table in db
        for index, row in tqdm(df.iterrows(), total=len(df)):
            columns = {
                self.db_verbose_names[col_name]: value for col_name, value in row.items()
            }
            try:
                self.model.objects.create(
                    id=index,
                    **columns
                )
                created_count += 1

            except Exception as e:
                err_count += 1
                pass

        end_time = time.time()
        mem_after = psutil.virtual_memory().used
        # print("DONE !!!")
        # print("Time for iterrows():", end_time - start_time, "seconds")
        # print("Memory used by iterrows():", mem_after - mem_before, "bytes")
        # print(f'created_count = {created_count}, err_count = {err_count}')
        return created_count, err_count

    def update_by_iterrows(self, df):
        # FOR TEST PURPOSES

        updated_count = 0
        err_count = 0

        start_time = time.time()
        mem_before = psutil.virtual_memory().used
        # col names in df are verbose names for table in db
        for index, row in tqdm(df.iterrows(), total=len(df)):
            columns = {
                self.db_verbose_names[col_name]: value for col_name, value in row.items()
            }
            try:
                # unique_identifier = index
                obj = self.model.objects.get(id=index)

                for col, value in columns.items():
                    setattr(obj, col, value)  # Dynamically set the attribute
                obj.save()  # Save the changes to the object
                updated_count += 1

            except Exception as e:
                # print(e)
                err_count += 1
                pass

        end_time = time.time()
        mem_after = psutil.virtual_memory().used
        # print("DONE !!!")
        # print("Time for iterrows():", end_time - start_time, "seconds")
        # print("Memory used by iterrows():", mem_after - mem_before, "bytes")
        # print(f'updated_count = {updated_count}, err_count = {err_count}')
        return updated_count, err_count

    def update_and_create_by_iterrows(self):
        # FOR TEST PURPOSES

        df, len_df = self.validate_df_to_get_unequal_rows()
        if df is not None:
            created_count, err_count_create = self.create_by_iterrows(df)
            df, len_df = self.validate_df_to_get_unequal_rows()
            updated_count, err_count_update = self.update_by_iterrows(df)
            return created_count, err_count_create, updated_count, err_count_update
        else:
            print('CHECK ERRORS!!! DB not updated!')
            return 0, 0, 0, 0

    # -----------------------
    def create_by_itertuples(self, df):
        created_count = 0
        err_count = 0

        start_time = time.time()
        mem_before = psutil.virtual_memory().used

        len_df = len(df)
        if len_df > 0:

            for row in tqdm(df.itertuples(), total=len_df):
                # col names in df are verbose names for table in db
                columns = {
                    self.col_list[i]: row[i+1] for i in range(0, len(self.col_list))
                }

                try:
                    # obj - for manual testing to look at obj's properties obj.art_number, ...
                    obj = self.model.objects.create(
                        id=row[0],
                        **columns,
                    )

                    created_count += 1

                except Exception as e:
                    # if err_count < 5:
                    #     print(row)
                    #     print(e)
                    err_count += 1
                    pass

        end_time = time.time()
        mem_after = psutil.virtual_memory().used
        # print("DONE !!!")
        # print("Time for itertuples():", end_time - start_time, "seconds")
        # print("Memory used by itertuples():", mem_after - mem_before, "bytes")
        # print(f'created_count = {created_count}, err_count = {err_count}')
        return created_count, err_count

    def update_by_itertuples(self, df, col_list):
        updated_count = 0
        err_count = 0

        start_time = time.time()
        mem_before = psutil.virtual_memory().used

        len_df = len(df)
        if len_df > 0:

            for row in tqdm(df.itertuples(), total=len_df):
                # col names in df are verbose names for table in db
                columns = {
                    col_list[i]: row[i + 1] for i in range(0, len(col_list))
                }
                try:
                    obj = self.model.objects.get(id=row[0])
                    for col, value in columns.items():
                        setattr(obj, col, value)  # Dynamically set the attribute

                    obj.save()  # Save the changes to the object
                    updated_count += 1

                except Exception as e:
                    # print(e)
                    err_count += 1
                    pass

        end_time = time.time()
        mem_after = psutil.virtual_memory().used
        # print("DONE !!!")
        # print("Time for itertuples():", end_time - start_time, "seconds")
        # print("Memory used by itertuples():", mem_after - mem_before, "bytes")
        # print(f'updated_count = {updated_count}, err_count = {err_count}')
        return updated_count, err_count

    def update_and_create_by_itertuples(self):
        df, len_df = self.validate_df_to_get_unequal_rows()
        if df is not None:
            created_count, err_count_create = self.create_by_itertuples(df)
            df, len_df = self.validate_df_to_get_unequal_rows()
            updated_count, err_count_update = self.update_by_itertuples(df, self.col_list)
            return created_count, err_count_create, updated_count, err_count_update
        else:
            print('CHECK ERRORS!!! DB not updated!')
            return 0, 0, 0, 0

    # When using iterrows(), NaN values are preserved in the resulting named tuple
    # When using itertuples(), NaN values are converted to str 'nan' (lowercase)
    # in the resulting named tuple. This is a pandas-specific NaN representation.

    def find_comment_colum_names(self):
        fields_of_comments_dict = {f'comments_to_field_{col_name}': self.db_column_names[col_name] for col_name in self.db_column_names.keys() if f"comments_to_field_{col_name}" in self.db_column_names.keys()}
        return fields_of_comments_dict

    def find_cell_names_to_comment_colum_names(self):
        cell_names = {cell_name: self.db_verbose_names[cell_value] for cell_name, cell_value in self.column_names_from_excel_file.items() if cell_value in self.fields_of_comments_dict.values()}
        return cell_names

    def read_db_comments_to_df(self):
        cell_names = self.find_cell_names_to_comment_colum_names()
        col_list = [f'comments_to_field_{col_name}' for k, col_name in cell_names.items()]
        df = self.read_specified_columns_from_db_table(col_list)

        return df, col_list

    def read_file_comments_to_df(self):
        empty_df, len_df = self.validate_df_to_get_unequal_rows()
        cell_names_with_comments = self.find_cell_names_to_comment_colum_names()  # dict
        comment_cell_names_dict = {
            cell_name: [
                cell_name[:-1] + str(x) for x in range(self.offset, len_df + self.offset)
            ] for cell_name in cell_names_with_comments.keys()
        }
        comment_dict = DataReader(
            self.path_to_file, self.file_name).read_comments_from_stock_file_by_openpyxl(
            self.sheet_name, comment_cell_names_dict)
        data = {
            self.db_verbose_names[file_col_name]: list_of_comments for file_col_name, list_of_comments in comment_dict.items()
        }
        data = {k.replace('_', ' '): v for k, v in data.items()}
        res_data = {('comments to field ' + str(k)): v for k, v in data.items()}

        df = pd.DataFrame(res_data)
        df.index = df.index + self.offset
        df = df.fillna(self.nan_value)

        return df

    def validate_comment_df_to_get_unequal_rows(self):
        db_comments, col_list = self.read_db_comments_to_df()
        file_comments = self.read_file_comments_to_df()
        comparison_result = db_comments == file_comments
        comments_to_update_df = file_comments[~comparison_result.all(axis=1)]
        fields = {verbose_name: verbose_name.replace(' ', '_') for verbose_name in comments_to_update_df.columns}
        comments_to_update_df = comments_to_update_df.rename(columns=fields)
        return comments_to_update_df, col_list

    def update_of_db(self):
        # rewrite later, now it works if input OK !!!!
        self.update_and_create_by_itertuples()
        comments_to_update_df, col_list = self.validate_comment_df_to_get_unequal_rows()
        updated_count, err_count = self.update_by_itertuples(comments_to_update_df, col_list)

